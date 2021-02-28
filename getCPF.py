import sys
import csv
import random
from openpyxl import Workbook
import string
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.comments import Comment

wb = Workbook()
sheet_cpf =wb.create_sheet(0)
sheet_cpf.title="CPF"

def cpf_funcional():

    n = [random.randrange(10) for i in range(9)]

    # calcula digito 1 e acrescenta ao numero
    s = sum(x * y for x, y in zip(n, range(10, 1, -1)))
    d1 = 11 - s % 11
    if d1 >= 10:
        d1 = 0
    n.append(d1)

    # calcula digito 2 e acrescenta ao numero
    s = sum(x * y for x, y in zip(n, range(11, 1, -1)))
    d2 = 11 - s % 11
    if d2 >= 10:
        d2 = 0
    n.append(d2)

    return "%d%d%d.%d%d%d.%d%d%d-%d%d" % tuple(n)

def gerar_cep():
    cep = [random.randint(0, 9) for x in range(8)]
    return '%s%s%s%s%s-%s%s%s' % tuple(cep)

i = 0
ccpf = 1
ccep = 1
position_cpf = 'B' + str(ccpf)
position_cep = 'C' + str(ccep)
cpf=['']*5001
cep=['']*5001

while i < 5000:
  cpf[i] = cpf_funcional()
  cep[i] = gerar_cep()
  sheet_cpf[position_cpf].value = cpf[i]
  sheet_cpf[position_cep].value = cep[i]
  i += 1
  ccpf += 1
  ccep += 1
  position_cpf = 'B' + str(ccpf)
  position_cep = 'C' + str(ccep)


wb.save('CPF.xlsx')
