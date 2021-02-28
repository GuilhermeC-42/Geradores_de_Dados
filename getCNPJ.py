import sys
import csv
import random
from openpyxl import Workbook
import string
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.comments import Comment

wb = Workbook()
sheet_cnpj =wb.create_sheet(0)
sheet_cnpj.title="CNPJ"


def cnpj_funcional():
    def calculate_special_digit(l):
        digit = 0

        for i, v in enumerate(l):
            digit += v * (i % 8 + 2)

        digit = 11 - digit % 11

        return digit if digit < 10 else 0

    cnpj =  [1, 0, 0, 0] + [random.randint(0, 9) for x in range(8)]

    for _ in range(2):
        cnpj = [calculate_special_digit(cnpj)] + cnpj

    return '%s%s.%s%s%s.%s%s%s/%s%s%s%s-%s%s' % tuple(cnpj[::-1])

def gerar_cep():
    cep = [random.randint(0, 9) for x in range(8)]
    return '%s%s%s%s%s-%s%s%s' % tuple(cep)

i = 0
ccnpj = 1
ccep = 1
position_cnpj = 'B' + str(ccnpj)
position_cep = 'C' + str(ccep)
cnpj=['']*5001
cep=['']*5001

while i < 5000:
  cnpj[i] = cnpj_funcional()
  cep[i] = gerar_cep()
  sheet_cnpj[position_cnpj].value = cnpj[i]
  sheet_cnpj[position_cep].value = cep[i]
  i += 1
  ccnpj += 1
  ccep += 1
  position_cnpj = 'B' + str(ccnpj)
  position_cep = 'C' + str(ccep)

wb.save('CNPJ.xlsx')
